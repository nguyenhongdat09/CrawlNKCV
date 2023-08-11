from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from Constant.Constant import Constant as ct
from selenium.common.exceptions import JavascriptException, NoSuchElementException, TimeoutException
import pandas as pd
import os
import datetime
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
pd.set_option('display.expand_frame_repr', False)  # Ngăn các cột bị cắt và xuống dòng

chromedriver_path = os.getcwd() + '/chromedriver.exe'
class loadWeb():
    def __init__(self):
        self.url = ct.url
        self.url_nkcv = ct.url_nkcv
        self.url_qlyc = ct.url_qlyc
    def init_driver(self):
        options = Options()
        options.add_experimental_option("detach", True)  # khong cho tat chorme
        options.add_argument("--start-maximized")  # Thêm tùy chọn để mở cửa sổ lớn nhất
        # service_object = Service(binary_path)
        dir = os.path.dirname(__file__)
        chrome_path = os.path.join(dir, 'selenium', 'webdriver', 'chromedriver_win32.exe')
        service_object = Service(chrome_path)
        driver = webdriver.Chrome(options = options, service=service_object)
        wait = WebDriverWait(driver, 10)
        return driver, wait

    def login(self):
        driver, wait = self.init_driver()
        while True:
            try:
                driver.get(self.url)
                UserName = driver.find_element(By.ID, "LoginExtender_txtUserName")
                UserName.send_keys(ct.user_name)
                password = driver.find_element(By.ID, "LoginExtender_txtPassword")
                password.send_keys(ct.password)
                button_login = driver.find_element(By.ID, "LoginExtender_Ok")
                button_login.click()
                wait.until(EC.presence_of_element_located((By.ID, "LoginExtender_Attention")) )
                WebDriverWait(driver.find_element(By.ID, "LoginExtender_Attention"), 3).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "Attention"))
                )
                js_script = "$find('LoginExtender')._login(true);"
                driver.execute_script(js_script)
                wait.until(EC.presence_of_element_located((By.CLASS_NAME, "MenuExtenderGroup")))
                # Lấy danh sách cookie từ tab thứ nhất
                cookies = driver.get_cookies()
                return driver, wait, cookies
            except TimeoutException:
                pass

    def enter_info_qlyc(self, driver, wait):
        driver.get(self.url_nkcv)

        wait.until(EC.presence_of_element_located((By.ID, "ctl00_FastBusiness_MainReport_searchExtender_form_ma_ltql")))

        input_ltql = driver.find_element(By.ID, "ctl00_FastBusiness_MainReport_searchExtender_form_ma_ltql")
        input_ltql.send_keys(ct.user_name)

        input_trang_thai = driver.find_element(By.ID, "ctl00_FastBusiness_MainReport_searchExtender_form_trang_thai")
        input_trang_thai.send_keys(ct.trang_thai)

        input_ngay_ht_tu = driver.find_element(By.ID, "ctl00_FastBusiness_MainReport_searchExtender_form_ngay_ht_tu")
        driver.execute_script("arguments[0].value = arguments[1];", input_ngay_ht_tu, ct.tu_ngay)

        input_ngay_ht_den = driver.find_element(By.ID, "ctl00_FastBusiness_MainReport_searchExtender_form_ngay_ht_den")
        driver.execute_script("arguments[0].value = arguments[1];", input_ngay_ht_den, ct.den_ngay)

        button_ok = driver.find_element(By.ID, "ctl00_FastBusiness_MainReport_searchExtender_updateDlgOk")
        button_ok.click()
        return self.crawl_info_qlyc(driver, wait)
    def set_size_page(self, driver, wait, size):
        scripts = (
            "var g = $find('ctl00_FastBusiness_MainReport');"
            f"g.set_gridPageSize({size});"
        )
        page_count = driver.execute_script(scripts)
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'td#ctl00_FastBusiness_MainReport_Wait a.PaddedLink')))
    def get_page_count(self, driver):
        scripts = (
            "var g = $find('ctl00_FastBusiness_MainReport');"
            "return g._gridPageCount "
        )
        return driver.execute_script(scripts)
    def go_to_page(self, driver, wait, page):
        scripts = (
            "var g = $find('ctl00_FastBusiness_MainReport');"
            f"g.goToPage('{page}'); "
        )
        driver.execute_script(scripts)
        wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'td#ctl00_FastBusiness_MainReport_Wait a.PaddedLink')))

    def crawl_info_qlyc(self, driver, wait):
        wait.until(EC.presence_of_element_located((By.ID, "ctl00_FastBusiness_MainReport_ToolbarButton_JobLog")))
        button_nkcv_hidden = driver.find_element(By.ID, 'ctl00_FastBusiness_MainReport_ToolbarButton_JobLog')
        button_nkcv_hidden.click()
        self.set_size_page(driver, wait, 250)
        page_count = self.get_page_count(driver)
        result = []
        for i in range(0, page_count):
            if i != 0:
                self.go_to_page(driver, wait, i)
            scripts = (
                "var g = $find('ctl00_FastBusiness_MainReport');"
                "var stt_rec_list = [], noi_dung_list = [], ngay_ht_list = [], ma_da_list = [];"
                "for (var i = 1; i <= g._rowCount; i++) {"
                "    ma_da_list.push(g._getItemValue(i, g._getColumnOrder('ma_da')));"
                "    stt_rec_list.push(g._getItemValue(i, g._getColumnOrder('stt_rec')));"
                "    noi_dung_list.push(g._getItemValue(i, g._getColumnOrder('noi_dung')));"
                "    ngay_ht_list.push(g._getItem(i, g._getColumnOrder('ngay_ht')).value);"
                "}"
                "return [stt_rec_list, ma_da_list, noi_dung_list, ngay_ht_list];"
            )
            result.append(driver.execute_script(scripts))
        df = pd.DataFrame(columns=["stt_rec", "ma_da", "noi_dung", "ngay_ht"])
        for item in result:
            dict = {
                "stt_rec": item[0],
                "ma_da": item[1],
                "noi_dung": item[2],
                "ngay_ht": item[3],
            }
            # Tạo DataFrame tạm thời từ dict data
            temp_df = pd.DataFrame(dict)
            df = pd.concat([df, temp_df], ignore_index=True)
        df['so_gio'] = None
        df['diem'] = None
        return df
    def enter_info_nkcv(self, driver, wait):
        driver.get(self.url_qlyc)

        wait.until(EC.presence_of_element_located((By.ID, "ctl00_FastBusiness_MainReport_searchExtender_updateDlgPanel")))
        input_ngay_tu = driver.find_element(By.ID,
                                                    "ctl00_FastBusiness_MainReport_searchExtender_form_ngay_tu")
        driver.execute_script("arguments[0].value = arguments[1];", input_ngay_tu, ct.tu_ngay)

        input_ngay_den = driver.find_element(By.ID,
                                                     "ctl00_FastBusiness_MainReport_searchExtender_form_ngay_den")
        driver.execute_script("arguments[0].value = arguments[1];", input_ngay_den, ct.den_ngay)
        button_ok = driver.find_element(By.ID, "ctl00_FastBusiness_MainReport_searchExtender_updateDlgOk")
        button_ok.click()
        return self.crawl_info_nkcv(driver, wait)
    def crawl_info_nkcv(self, driver, wait):
        wait.until(
            EC.presence_of_element_located((By.ID, "ctl00_FastBusiness_MainReport_ToolbarButton_New")))
        result = []
        self.set_size_page(driver, wait, 25)
        page_count = self.get_page_count(driver)
        for i in range(0, page_count):
            if i != 0:
                self.go_to_page(driver, wait, i)
            scripts = (
                "var g = $find('ctl00_FastBusiness_MainReport');"
                "var  stt_rec_list = [], noi_dung_list = [], so_gio_list = [], ngay_list = [];"
                "for (var i = 1; i <= g._rowCount; i++){"
                "   stt_rec_list.push(g._getItemValue(i, g._getColumnOrder('stt_rec')));"
                "   noi_dung_list.push(g._getItemValue(i, g._getColumnOrder('noi_dung')));"
                "   so_gio_list.push(g._getItemValue(i, g._getColumnOrder('so_gio')));"
                "   ngay_list.push(g._getItem(i, g._getColumnOrder('ngay_ct')).value);"
                "}"
                "return [stt_rec_list, ngay_list, noi_dung_list, so_gio_list];"
            )
            result.append(driver.execute_script(scripts))

        df = pd.DataFrame(columns=['stt_rec', 'ngay_ht', 'noi_dung', 'so_gio'])
        for item in result:
            dict = {
                "stt_rec": item[0],
                "ngay_ht": item[1],
                "noi_dung": item[2],
                "so_gio": item[3],
            }
            temp_df = pd.DataFrame(dict)
            df = pd.concat([df, temp_df], ignore_index= True)
        return df


    def run(self):
        driver, wait, cookies = self.login()
        df_qlyc = self.enter_info_qlyc(driver, wait)
        df_nkcv = self.enter_info_nkcv(driver, wait)
        gio_da_nhap = sum(df_nkcv['so_gio'])
        self.hour_calculate(gio_da_nhap)
        self.export_excel(df_qlyc, df_nkcv)
        driver.quit()
    def export_excel(self, df_qlyc, df_nkcv):
        # Tạo một workbook mới
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Nhat ky chua nhap"  # Đặt tên cho sheet 1
        ws2 = wb.create_sheet(title="Nhat ky da nhap")  # Tạo sheet 2
        df1 = pd.concat([df_qlyc, df_nkcv], ignore_index=True)
        df1 = self.sort_df(df1)
        self.write_dataframe_to_sheet(df1, ws1, 3, 110)
        self.write_dataframe_to_sheet(df_nkcv, ws2, 3, 110)
        # Lưu workbook vào file Excel
        wb.save(ct.ten_file_excel)
    def sort_df(self, df):
        df["ngay_ht"] = pd.to_datetime(df["ngay_ht"], format="%d/%m/%Y")
        df = df.sort_values(by="ngay_ht")
        df["ngay_ht"] = pd.to_datetime(df["ngay_ht"]).dt.strftime('%d/%m/%Y')

        return df
    # Hàm để ghi DataFrame vào sheet
    def write_dataframe_to_sheet(self, df, sheet, col_index, width):
        for i, column in enumerate(df.columns, 1):
            if i == col_index:
                sheet.column_dimensions[chr(64 + i)].width = width
            else:
                sheet.column_dimensions[chr(64 + i)].width = 16
        for row in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row)
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(wrapText=True)
    def hour_calculate(self, gio_da_nhap):
        tu_ngay, den_ngay = datetime.datetime.strptime(ct.tu_ngay, '%d/%m/%Y'), datetime.datetime.strptime(ct.den_ngay, '%d/%m/%Y')
        df = pd.DataFrame(columns=['ngay', 'gio'])
        while tu_ngay <= den_ngay:
            if tu_ngay.weekday() >= 0 and tu_ngay.weekday() <= 5:
                dict ={
                    "ngay": tu_ngay
                }
                if tu_ngay.weekday() == 5:
                    dict['gio'] = 4
                else:
                    dict['gio'] = 8
                temp_df = pd.DataFrame([dict])
                df = pd.concat([df, temp_df], ignore_index=True)
            tu_ngay += datetime.timedelta(days=1)
        total_hours = sum(df['gio'])
        print(f'Từ ngày {ct.tu_ngay} đến ngày {ct.den_ngay} bạn phải nhập "{total_hours}" giờ. Hãy mở file excel "{ct.ten_file_excel}" nhập đủ "{total_hours}" giờ, sau đó tôi sẽ bổ giờ theo ngày tự động cho bạn')
        return df

# crawl = loadWeb()
# crawl.run()
